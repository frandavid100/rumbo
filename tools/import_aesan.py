#!/usr/bin/env python3
"""Build the bundled food catalogue from AESAN's open 2022 workbook."""

from __future__ import annotations

import argparse
import gzip
import json
import re
import unicodedata
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


SOURCE_URL = (
    "https://www.aesan.gob.es/AECOSAN/docs/documentos/seguridad_alimentaria/"
    "bases_datos/BasedatosWeb.xlsx"
)
SOURCE_PAGE = (
    "https://www.aesan.gob.es/AECOSAN/web/seguridad_alimentaria/"
    "subseccion/alimentosBebidas.htm"
)
AESAN_ID_OFFSET = 100_000_000_000_000


def normalized(value: object | None) -> str:
    text = "" if value is None else str(value)
    return "".join(
        char for char in unicodedata.normalize("NFD", text.lower())
        if unicodedata.category(char) != "Mn"
    )


def nutritional_role(category: str, subcategory: str, name: str) -> str:
    text = normalized(f"{category} {subcategory} {name}")
    if re.search(r"\b(fruta|frutas|zumo|zumos|nectar|mosto)\b", text) and not re.search(
        r"frutos secos|barrita|pastel|galleta|helado", text
    ):
        return "FRUIT"
    if re.search(r"hortaliz|verdura|legumbre", text) and not re.search(
        r"patatas fritas|aperitivo|plato preparado", text
    ):
        return "VEGETABLE"
    if re.search(
        r"carne|aves|pescado|huevo|queso|yogur|leche fermentada|analogo de carne", text
    ):
        return "PROTEIN"
    if re.search(r"mantequilla|grasa|aceite|mayonesa|frutos secos", text):
        return "FAT"
    if re.search(r"pan|pasta|arroz|grano|cereal|bolleria|pastel|galleta", text):
        return "CARBOHYDRATE"
    return "OTHER"


def number(value: object | None) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def text(value: object | None) -> str | None:
    if value is None:
        return None
    result = re.sub(r"\s+", " ", str(value)).strip()
    return result or None


def barcode(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def download_if_needed(workbook: Path) -> None:
    if workbook.exists():
        return
    workbook.parent.mkdir(parents=True, exist_ok=True)
    request = urllib.request.Request(SOURCE_URL, headers={"User-Agent": "Rumbo catalogue builder"})
    with urllib.request.urlopen(request) as response, workbook.open("wb") as target:
        target.write(response.read())


def build(workbook: Path, output: Path) -> tuple[int, int]:
    book = load_workbook(workbook, read_only=True, data_only=True)
    sheet = book["Tabla1"]
    header = [cell.value for cell in next(sheet.iter_rows())]
    column = {name: index for index, name in enumerate(header)}
    required = {
        "EAN", "Nombrecomercial", "Categoria", "Subcategoria", "Marca", "Submarca",
        "Fabricante", "DenominacionLegal", "Ingredientes", "EnergiaKC", "Grasas",
        "GrasasSat", "Carbohidratos", "Azúcares", "Proteínas", "Sal", "Fibra"
    }
    missing = required - column.keys()
    if missing:
        raise RuntimeError(f"The AESAN workbook is missing columns: {sorted(missing)}")

    output.parent.mkdir(parents=True, exist_ok=True)
    total = complete = 0
    seen_barcodes: set[str] = set()
    with gzip.open(output, "wt", encoding="utf-8", newline="\n", compresslevel=9) as target:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ean_value = row[column["EAN"]]
            name = text(row[column["Nombrecomercial"]])
            if ean_value is None or name is None:
                continue
            ean = barcode(ean_value)
            if ean in seen_barcodes:
                continue
            seen_barcodes.add(ean)
            family = text(row[column["Categoria"]])
            subcategory = text(row[column["Subcategoria"]])
            manufacturer = normalized(row[column["Fabricante"]])
            brand = text(row[column["Marca"]])
            subbrand = normalized(row[column["Submarca"]])
            retailer = (
                "Mercadona" if "mercadona" in manufacturer or "hacendado" in subbrand or
                "hacendado" in normalized(brand) else None
            )
            nutrients = {
                "k": number(row[column["EnergiaKC"]]),
                "f": number(row[column["Grasas"]]),
                "c": number(row[column["Carbohidratos"]]),
                "p": number(row[column["Proteínas"]]),
            }
            if all(value is not None for value in nutrients.values()):
                complete += 1
            item = {
                "i": AESAN_ID_OFFSET + int(ean),
                "n": name,
                "r": nutritional_role(family or "", subcategory or "", name),
                **nutrients,
                "fi": number(row[column["Fibra"]]),
                "b": ean,
                "br": brand,
                "fa": family,
                "sc": subcategory,
                "ln": text(row[column["DenominacionLegal"]]),
                "ing": text(row[column["Ingredientes"]]),
                "sat": number(row[column["GrasasSat"]]),
                "su": number(row[column["Azúcares"]]),
                "sa": number(row[column["Sal"]]),
                "ret": retailer,
            }
            target.write(json.dumps(item, ensure_ascii=False, separators=(",", ":")) + "\n")
            total += 1
    return total, complete


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=Path("build/aesan/BasedatosWeb.xlsx"))
    parser.add_argument(
        "--output", type=Path,
        default=Path("app/src/main/assets/aesan_foods.dat")
    )
    args = parser.parse_args()
    download_if_needed(args.workbook)
    total, complete = build(args.workbook, args.output)
    print(f"Imported {total} products ({complete} with complete core nutrition) from {SOURCE_PAGE}")


if __name__ == "__main__":
    main()
